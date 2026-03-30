#!/usr/bin/env python3
"""
EV Charging Report Generator
=============================
Generuje mesacne a rocne reporty zo suboru ev_charging_sessions.csv

Pouzitie:
  python3 ev_charging_report.py              # Aktualny mesiac
  python3 ev_charging_report.py 2025-01      # Konkretny mesiac
  python3 ev_charging_report.py --year 2025  # Rocny report
  python3 ev_charging_report.py --all        # Vsetky data

Vystup:
  - CSV suhrn
  - XLSX report (ak je nainstalovany openpyxl)
"""

import csv
import os
import sys
from datetime import datetime
from collections import defaultdict

# Cesta k CSV suboru
CSV_FILE = "/config/ev_charging_sessions.csv"
OUTPUT_DIR = "/config/ev_reports/"
WWW_DIR = "/config/www/ev_reports/"

def ensure_dirs():
    """Vytvori output adresare ak neexistuju"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(WWW_DIR, exist_ok=True)

def load_sessions(csv_file):
    """Nacita sessions z CSV suboru"""
    sessions = []
    
    if not os.path.exists(csv_file):
        print(f"Subor {csv_file} neexistuje!")
        return sessions
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                sessions.append({
                    'session_id': row.get('session_id', ''),
                    'start_time': datetime.strptime(row.get('start_time', ''), '%Y-%m-%d %H:%M:%S'),
                    'end_time': datetime.strptime(row.get('end_time', ''), '%Y-%m-%d %H:%M:%S'),
                    'user': row.get('user', 'unknown'),
                    'energy_kwh': float(row.get('energy_kwh', 0)),
                    'cost_eur': float(row.get('cost_eur', 0)),
                    'tariff': row.get('tariff', ''),
                    'stop_reason': row.get('stop_reason', ''),
                    'duration_min': float(row.get('duration_min', 0))
                })
            except (ValueError, KeyError) as e:
                print(f"Chyba pri parsovani riadku: {e}")
                continue
    
    return sessions

def filter_by_month(sessions, year_month):
    """Filtruje sessions podla mesiaca (format: YYYY-MM)"""
    year, month = map(int, year_month.split('-'))
    return [s for s in sessions if s['start_time'].year == year and s['start_time'].month == month]

def filter_by_year(sessions, year):
    """Filtruje sessions podla roku"""
    return [s for s in sessions if s['start_time'].year == int(year)]

def generate_report(sessions, title="EV Charging Report"):
    """Generuje report zo sessions"""
    
    if not sessions:
        return {
            'title': title,
            'total_sessions': 0,
            'total_energy_kwh': 0,
            'total_cost_eur': 0,
            'total_duration_min': 0,
            'avg_cost_per_kwh': 0,
            'by_user': {},
            'by_tariff': {},
            'sessions': []
        }
    
    # Agregacie
    by_user = defaultdict(lambda: {'sessions': 0, 'energy': 0, 'cost': 0, 'duration': 0})
    by_tariff = defaultdict(lambda: {'sessions': 0, 'energy': 0, 'cost': 0})
    
    total_energy = 0
    total_cost = 0
    total_duration = 0
    
    for s in sessions:
        total_energy += s['energy_kwh']
        total_cost += s['cost_eur']
        total_duration += s['duration_min']
        
        by_user[s['user']]['sessions'] += 1
        by_user[s['user']]['energy'] += s['energy_kwh']
        by_user[s['user']]['cost'] += s['cost_eur']
        by_user[s['user']]['duration'] += s['duration_min']
        
        by_tariff[s['tariff']]['sessions'] += 1
        by_tariff[s['tariff']]['energy'] += s['energy_kwh']
        by_tariff[s['tariff']]['cost'] += s['cost_eur']
    
    return {
        'title': title,
        'total_sessions': len(sessions),
        'total_energy_kwh': round(total_energy, 2),
        'total_cost_eur': round(total_cost, 2),
        'total_duration_min': round(total_duration, 0),
        'avg_cost_per_kwh': round(total_cost / total_energy, 4) if total_energy > 0 else 0,
        'by_user': dict(by_user),
        'by_tariff': dict(by_tariff),
        'sessions': sessions
    }

def save_csv_report(report, filename):
    """Ulozi report do CSV"""
    with open(filename, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        
        writer.writerow(['EV Charging Report'])
        writer.writerow([report['title']])
        writer.writerow([])
        
        writer.writerow(['SUHRN'])
        writer.writerow(['Pocet sessions', report['total_sessions']])
        writer.writerow(['Celkova energia (kWh)', report['total_energy_kwh']])
        writer.writerow(['Celkove naklady (EUR)', report['total_cost_eur']])
        writer.writerow(['Celkovy cas (min)', report['total_duration_min']])
        writer.writerow(['Priemerna cena (EUR/kWh)', report['avg_cost_per_kwh']])
        writer.writerow([])
        
        writer.writerow(['PODLA POUZIVATELA'])
        writer.writerow(['Pouzivatel', 'Sessions', 'Energia (kWh)', 'Naklady (EUR)', 'Cas (min)'])
        for user, data in report['by_user'].items():
            writer.writerow([user, data['sessions'], round(data['energy'], 2), round(data['cost'], 2), round(data['duration'], 0)])
        writer.writerow([])
        
        writer.writerow(['PODLA TARIFY'])
        writer.writerow(['Tarifa', 'Sessions', 'Energia (kWh)', 'Naklady (EUR)'])
        for tariff, data in report['by_tariff'].items():
            writer.writerow([tariff, data['sessions'], round(data['energy'], 2), round(data['cost'], 2)])
        writer.writerow([])
        
        if report['sessions']:
            writer.writerow(['DETAIL SESSIONS'])
            writer.writerow(['Datum', 'Pouzivatel', 'Energia (kWh)', 'Naklady (EUR)', 'Tarifa', 'Trvanie (min)'])
            for s in report['sessions']:
                writer.writerow([
                    s['start_time'].strftime('%Y-%m-%d %H:%M'),
                    s['user'],
                    round(s['energy_kwh'], 2),
                    round(s['cost_eur'], 2),
                    s['tariff'],
                    round(s['duration_min'], 0)
                ])
    
    print(f"CSV report ulozeny: {filename}")

def save_xlsx_report(report, filename):
    """Ulozi report do XLSX (vyzaduje openpyxl)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl nie je nainstalovany. XLSX report preskoceny.")
        print("Pre instalaciu: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    ws.cell(row=row, column=1, value="EV Charging Report").font = header_font
    row += 1
    ws.cell(row=row, column=1, value=report['title']).font = subheader_font
    row += 2
    
    ws.cell(row=row, column=1, value="SUHRN").font = subheader_font
    row += 1
    
    summary_data = [
        ('Pocet sessions', report['total_sessions']),
        ('Celkova energia (kWh)', report['total_energy_kwh']),
        ('Celkove naklady (EUR)', report['total_cost_eur']),
        ('Celkovy cas (min)', report['total_duration_min']),
        ('Priemerna cena (EUR/kWh)', report['avg_cost_per_kwh'])
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    
    ws.cell(row=row, column=1, value="PODLA POUZIVATELA").font = subheader_font
    row += 1
    
    headers = ['Pouzivatel', 'Sessions', 'Energia (kWh)', 'Naklady (EUR)', 'Cas (min)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    row += 1
    
    for user, data in report['by_user'].items():
        ws.cell(row=row, column=1, value=user).border = thin_border
        ws.cell(row=row, column=2, value=data['sessions']).border = thin_border
        ws.cell(row=row, column=3, value=round(data['energy'], 2)).border = thin_border
        ws.cell(row=row, column=4, value=round(data['cost'], 2)).border = thin_border
        ws.cell(row=row, column=5, value=round(data['duration'], 0)).border = thin_border
        row += 1
    
    row += 1
    
    ws.cell(row=row, column=1, value="PODLA TARIFY").font = subheader_font
    row += 1
    
    headers = ['Tarifa', 'Sessions', 'Energia (kWh)', 'Naklady (EUR)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    row += 1
    
    for tariff, data in report['by_tariff'].items():
        ws.cell(row=row, column=1, value=tariff).border = thin_border
        ws.cell(row=row, column=2, value=data['sessions']).border = thin_border
        ws.cell(row=row, column=3, value=round(data['energy'], 2)).border = thin_border
        ws.cell(row=row, column=4, value=round(data['cost'], 2)).border = thin_border
        row += 1
    
    row += 1
    
    if report['sessions']:
        ws.cell(row=row, column=1, value="DETAIL SESSIONS").font = subheader_font
        row += 1
        
        headers = ['Datum', 'Pouzivatel', 'Energia (kWh)', 'Naklady (EUR)', 'Tarifa', 'Trvanie (min)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = thin_border
        row += 1
        
        for s in report['sessions']:
            ws.cell(row=row, column=1, value=s['start_time'].strftime('%Y-%m-%d %H:%M')).border = thin_border
            ws.cell(row=row, column=2, value=s['user']).border = thin_border
            ws.cell(row=row, column=3, value=round(s['energy_kwh'], 2)).border = thin_border
            ws.cell(row=row, column=4, value=round(s['cost_eur'], 2)).border = thin_border
            ws.cell(row=row, column=5, value=s['tariff']).border = thin_border
            ws.cell(row=row, column=6, value=round(s['duration_min'], 0)).border = thin_border
            row += 1
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    wb.save(filename)
    print(f"XLSX report ulozeny: {filename}")

def main():
    ensure_dirs()
    
    sessions = load_sessions(CSV_FILE)
    
    if not sessions:
        print("Ziadne sessions na spracovanie.")
        return
    
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        
        if arg == '--all':
            report = generate_report(sessions, "Vsetky sessions")
            base_name = "ev_report_all"
        elif arg == '--year' and len(sys.argv) > 2:
            year = sys.argv[2]
            filtered = filter_by_year(sessions, year)
            report = generate_report(filtered, f"Rocny report {year}")
            base_name = f"ev_report_{year}"
        elif '-' in arg and len(arg) == 7:
            filtered = filter_by_month(sessions, arg)
            report = generate_report(filtered, f"Mesacny report {arg}")
            base_name = f"ev_report_{arg}"
        else:
            print(f"Neznamy argument: {arg}")
            print("Pouzitie: python3 ev_charging_report.py [YYYY-MM | --year YYYY | --all]")
            return
    else:
        current_month = datetime.now().strftime('%Y-%m')
        filtered = filter_by_month(sessions, current_month)
        report = generate_report(filtered, f"Mesacny report {current_month}")
        base_name = f"ev_report_{current_month}"
    
    csv_file = os.path.join(OUTPUT_DIR, f"{base_name}.csv")
    xlsx_file = os.path.join(OUTPUT_DIR, f"{base_name}.xlsx")
    
    save_csv_report(report, csv_file)
    save_xlsx_report(report, xlsx_file)
    
    import shutil
    shutil.copy(csv_file, os.path.join(WWW_DIR, f"{base_name}.csv"))
    if os.path.exists(xlsx_file):
        shutil.copy(xlsx_file, os.path.join(WWW_DIR, f"{base_name}.xlsx"))
    
    print(f"\n{'='*50}")
    print(f"REPORT: {report['title']}")
    print(f"{'='*50}")
    print(f"Sessions: {report['total_sessions']}")
    print(f"Energia: {report['total_energy_kwh']} kWh")
    print(f"Naklady: {report['total_cost_eur']} EUR")
    print(f"Priemerna cena: {report['avg_cost_per_kwh']} EUR/kWh")
    print(f"{'='*50}")

if __name__ == "__main__":
    main()
