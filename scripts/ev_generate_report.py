#!/usr/bin/env python3
"""
EV Charging Report Generator
Funguje s 9-stlpcovym aj 12-stlpcovym CSV.
Pouzitie:
  python3 ev_generate_report.py xlsx
  python3 ev_generate_report.py monthly
  python3 ev_generate_report.py yearly
  python3 ev_generate_report.py refresh
"""

import csv, sys, os
from datetime import datetime
from collections import defaultdict

CSV_PATH = "/config/ev_charging_sessions.csv"
WWW_DIR = "/config/www"
REPORT_DIR = "/config/www/ev_reports"

# Hlavicky - 9 zakladnych
FIELDS_9 = ['session_id','start_time','end_time','user','energy_kwh',
            'cost_eur','tariff','stop_reason','duration_min']
# 12 s prebytkami
FIELDS_12 = FIELDS_9 + ['surplus_energy_kwh','grid_energy_kwh','surplus_value_eur']

def sf(val, default=0):
    try: return float(val)
    except: return default

def read_csv():
    sessions = []
    if not os.path.exists(CSV_PATH):
        return sessions
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        first_line = f.readline().strip()
        f.seek(0)
        
        # Detekuj pocet stlpcov
        if 'surplus' in first_line:
            fieldnames = FIELDS_12
        elif first_line.startswith('session_id'):
            fieldnames = None  # pouzije hlavicku z CSV
        else:
            fieldnames = FIELDS_9
        
        if fieldnames:
            reader = csv.DictReader(f, fieldnames=fieldnames)
            # Preskoc hlavicku ak existuje
            first = next(reader, None)
            if first and first.get('session_id','').startswith('session_id'):
                pass  # bola to hlavicka, preskocena
            elif first:
                sessions.append(first)
        else:
            reader = csv.DictReader(f)
        
        for row in reader:
            try:
                row['energy_kwh'] = sf(row.get('energy_kwh'))
                row['cost_eur'] = sf(row.get('cost_eur'))
                row['duration_min'] = sf(row.get('duration_min'))
                row['surplus_energy_kwh'] = sf(row.get('surplus_energy_kwh'))
                row['grid_energy_kwh'] = sf(row.get('grid_energy_kwh'))
                row['surplus_value_eur'] = sf(row.get('surplus_value_eur'))
                
                start_str = row.get('start_time', '').strip()
                row['start_dt'] = None
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M']:
                    try:
                        row['start_dt'] = datetime.strptime(start_str, fmt)
                        break
                    except: pass
                sessions.append(row)
            except Exception as e:
                print(f"Skip row: {e}", file=sys.stderr)
    return sessions

def generate_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    sessions = read_csv()
    if not sessions:
        print("Ziadne data")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions"
    
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    brd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    
    headers = ['Datum','Cas','Pouzivatel','Energia (kWh)','Naklady (EUR)',
               'Tarifa','Trvanie (min)','Dovod',
               'Prebytky (kWh)','Siet (kWh)','Uspora (EUR)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = brd
    
    sfill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    for ri, s in enumerate(sessions, 2):
        dt = s.get('start_dt')
        vals = [
            dt.strftime('%Y-%m-%d') if dt else s.get('start_time',''),
            dt.strftime('%H:%M') if dt else '',
            s.get('user',''),
            s['energy_kwh'],
            s['cost_eur'],
            s.get('tariff',''),
            s['duration_min'],
            s.get('stop_reason',''),
            s['surplus_energy_kwh'],
            s['grid_energy_kwh'],
            s['surplus_value_eur']
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = brd
            if ci in (4, 5, 9, 10, 11):
                cell.number_format = '0.000'
            if ci in (9, 11) and isinstance(v, (int, float)) and v > 0:
                cell.fill = sfill
    
    # Sucet riadok
    sr = len(sessions) + 2
    ws.cell(row=sr, column=3, value="SPOLU:").font = Font(bold=True)
    for ci, k in [(4,'energy_kwh'), (5,'cost_eur'), (7,'duration_min'),
                   (9,'surplus_energy_kwh'), (10,'grid_energy_kwh'), (11,'surplus_value_eur')]:
        c = ws.cell(row=sr, column=ci, value=round(sum(s[k] for s in sessions), 3))
        c.font = Font(bold=True)
    
    # Sirky stlpcov
    for i, w in enumerate([12, 8, 14, 14, 14, 8, 12, 16, 14, 14, 14]):
        ws.column_dimensions[chr(65+i)].width = w
    
    # Sheet 2: Mesacny prehlad
    ws2 = wb.create_sheet("Mesacny prehlad")
    mh = ['Mesiac','Sessions','Energia (kWh)','Naklady (EUR)','Trvanie (min)']
    for c, h in enumerate(mh, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
    by_month = defaultdict(lambda: {'n':0, 'e':0, 'c':0, 'd':0})
    for s in sessions:
        if s.get('start_dt'):
            mk = s['start_dt'].strftime('%Y-%m')
            by_month[mk]['n'] += 1
            by_month[mk]['e'] += s['energy_kwh']
            by_month[mk]['c'] += s['cost_eur']
            by_month[mk]['d'] += s['duration_min']
    for ri, (m, d) in enumerate(sorted(by_month.items()), 2):
        for ci, v in enumerate([m, d['n'], round(d['e'],3), round(d['c'],3), round(d['d'],0)], 1):
            ws2.cell(row=ri, column=ci, value=v)
    
    # Sheet 3: Pouzivatelia
    ws3 = wb.create_sheet("Pouzivatelia")
    uh = ['Pouzivatel','Sessions','Energia (kWh)','Naklady (EUR)']
    for c, h in enumerate(uh, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
    by_user = defaultdict(lambda: {'n':0, 'e':0, 'c':0})
    for s in sessions:
        u = s.get('user','?')
        by_user[u]['n'] += 1
        by_user[u]['e'] += s['energy_kwh']
        by_user[u]['c'] += s['cost_eur']
    for ri, (u, d) in enumerate(sorted(by_user.items()), 2):
        for ci, v in enumerate([u, d['n'], round(d['e'],3), round(d['c'],3)], 1):
            ws3.cell(row=ri, column=ci, value=v)
    
    os.makedirs(WWW_DIR, exist_ok=True)
    os.makedirs(REPORT_DIR, exist_ok=True)
    
    p = os.path.join(WWW_DIR, "ev_charging_sessions.xlsx")
    wb.save(p)
    os.chmod(p, 0o644)
    
    month_name = datetime.now().strftime('%Y-%m')
    p2 = os.path.join(REPORT_DIR, f"ev_report_{month_name}.xlsx")
    wb.save(p2)
    os.chmod(p2, 0o644)
    
    print(f"XLSX: {p} ({len(sessions)} sessions)")

def refresh_csv():
    if os.path.exists(CSV_PATH):
        os.makedirs(WWW_DIR, exist_ok=True)
        os.system(f"cp '{CSV_PATH}' '{WWW_DIR}/ev_charging_sessions.csv'")
        print("CSV refreshed")

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "xlsx"
    if cmd == "xlsx":
        generate_xlsx()
    elif cmd == "refresh":
        refresh_csv()
    elif cmd == "monthly":
        generate_xlsx()  # xlsx obsahuje mesacny prehlad
    elif cmd == "yearly":
        generate_xlsx()  # xlsx obsahuje vsetko
    else:
        print(f"Pouzitie: xlsx|refresh|monthly|yearly")