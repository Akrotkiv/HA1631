#!/usr/bin/env python3
"""
EV Charging Report Generator v2
CSV format (12 stlpcov): session_id,start_time,end_time,user,energy_kwh,
  cost_eur,tariff,stop_reason,duration_min,surplus_energy_kwh,
  grid_energy_kwh,surplus_value_eur

Pouzitie:
  python3 /config/scripts/ev_generate_report.py monthly [YYYY-MM]
  python3 /config/scripts/ev_generate_report.py yearly [YYYY]
  python3 /config/scripts/ev_generate_report.py xlsx
  python3 /config/scripts/ev_generate_report.py refresh
"""

import csv, sys, os, json
from datetime import datetime
from collections import defaultdict

CSV_PATH = "/config/ev_charging_sessions.csv"
WWW_DIR = "/config/www"
REPORT_DIR = "/config/www/ev_reports"

FIELDNAMES = [
    'session_id', 'start_time', 'end_time', 'user', 'energy_kwh',
    'cost_eur', 'tariff', 'stop_reason', 'duration_min',
    'surplus_energy_kwh', 'grid_energy_kwh', 'surplus_value_eur'
]

def ensure_dirs():
    os.makedirs(WWW_DIR, exist_ok=True)
    os.makedirs(REPORT_DIR, exist_ok=True)

def sf(val, default=0):
    try: return float(val)
    except: return default

def read_csv():
    sessions = []
    if not os.path.exists(CSV_PATH):
        return sessions
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        first = f.readline().strip()
        f.seek(0)
        if first.startswith('session_id'):
            reader = csv.DictReader(f)
        else:
            reader = csv.DictReader(f, fieldnames=FIELDNAMES)
        for row in reader:
            try:
                for k in ['energy_kwh','cost_eur','duration_min','surplus_energy_kwh','grid_energy_kwh','surplus_value_eur']:
                    row[k] = sf(row.get(k))
                start_str = row.get('start_time', '').strip()
                row['start_dt'] = None
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S']:
                    try:
                        row['start_dt'] = datetime.strptime(start_str, fmt)
                        break
                    except: pass
                sessions.append(row)
            except Exception as e:
                print(f"Skip: {e}", file=sys.stderr)
    return sessions

def generate_monthly_report(year_month=None):
    sessions = read_csv()
    if not sessions:
        print("Ziadne data"); return
    if year_month is None:
        year_month = datetime.now().strftime('%Y-%m')
    y, m = int(year_month.split('-')[0]), int(year_month.split('-')[1])
    monthly = [s for s in sessions if s.get('start_dt') and s['start_dt'].year == y and s['start_dt'].month == m]
    
    te = sum(s['energy_kwh'] for s in monthly)
    tc = sum(s['cost_eur'] for s in monthly)
    td = sum(s['duration_min'] for s in monthly)
    ts = sum(s['surplus_energy_kwh'] for s in monthly)
    tg = sum(s['grid_energy_kwh'] for s in monthly)
    tv = sum(s['surplus_value_eur'] for s in monthly)
    
    by_user = defaultdict(lambda: {k:0 for k in ['n','e','c','d','s','g']})
    by_tariff = defaultdict(lambda: {k:0 for k in ['n','e','c','s']})
    for s in monthly:
        u, t = s.get('user','?'), s.get('tariff','?')
        by_user[u]['n']+=1; by_user[u]['e']+=s['energy_kwh']; by_user[u]['c']+=s['cost_eur']
        by_user[u]['d']+=s['duration_min']; by_user[u]['s']+=s['surplus_energy_kwh']; by_user[u]['g']+=s['grid_energy_kwh']
        by_tariff[t]['n']+=1; by_tariff[t]['e']+=s['energy_kwh']; by_tariff[t]['c']+=s['cost_eur']; by_tariff[t]['s']+=s['surplus_energy_kwh']
    
    ensure_dirs()
    txt = os.path.join(REPORT_DIR, f"ev_report_{year_month}.txt")
    with open(txt, 'w', encoding='utf-8') as f:
        f.write(f"{'='*45}\n  EV NABIJANIE - MESACNY REPORT {year_month}\n{'='*45}\n\n")
        f.write(f"  Sessions:    {len(monthly)}\n")
        f.write(f"  Energia:     {te:.3f} kWh\n")
        f.write(f"    prebytky:  {ts:.3f} kWh ({ts/max(te,0.001)*100:.1f}%)\n")
        f.write(f"    siet:      {tg:.3f} kWh\n")
        f.write(f"  Naklady:     {tc:.3f} EUR (len siet)\n")
        f.write(f"  Uspora:      {tv:.3f} EUR (prebytky)\n")
        f.write(f"  Cas:         {td:.0f} min ({td/60:.1f} h)\n")
        if tg > 0: f.write(f"  Cena/kWh:    {tc/tg:.5f} EUR\n")
        f.write(f"\n--- Pouzivatelia ---\n")
        for u, d in sorted(by_user.items()):
            f.write(f"  {u}: {d['n']}x, {d['e']:.3f} kWh (prebytky: {d['s']:.3f}), {d['c']:.3f} EUR\n")
        f.write(f"\n--- Tarify ---\n")
        for t, d in sorted(by_tariff.items()):
            f.write(f"  {t}: {d['n']}x, {d['e']:.3f} kWh (prebytky: {d['s']:.3f}), {d['c']:.3f} EUR\n")
    
    os.system(f"cp '{txt}' '{WWW_DIR}/ev_monthly_report.txt'; chmod 644 '{WWW_DIR}/ev_monthly_report.txt' '{txt}'")
    print(f"Monthly: {len(monthly)} sessions, {te:.3f} kWh, {tc:.3f} EUR, surplus: {ts:.3f} kWh")

def generate_yearly_report(year=None):
    sessions = read_csv()
    if not sessions:
        print("Ziadne data"); return
    year = int(year) if year else datetime.now().year
    yearly = [s for s in sessions if s.get('start_dt') and s['start_dt'].year == year]
    
    by_month = defaultdict(lambda: {k:0 for k in ['n','e','c','s','g','v','d']})
    by_user = defaultdict(lambda: {k:0 for k in ['n','e','c','s']})
    for s in yearly:
        mk = s['start_dt'].strftime('%Y-%m')
        u = s.get('user','?')
        by_month[mk]['n']+=1; by_month[mk]['e']+=s['energy_kwh']; by_month[mk]['c']+=s['cost_eur']
        by_month[mk]['s']+=s['surplus_energy_kwh']; by_month[mk]['g']+=s['grid_energy_kwh']
        by_month[mk]['v']+=s['surplus_value_eur']; by_month[mk]['d']+=s['duration_min']
        by_user[u]['n']+=1; by_user[u]['e']+=s['energy_kwh']; by_user[u]['c']+=s['cost_eur']; by_user[u]['s']+=s['surplus_energy_kwh']
    
    te=sum(s['energy_kwh'] for s in yearly); tc=sum(s['cost_eur'] for s in yearly)
    ts=sum(s['surplus_energy_kwh'] for s in yearly); tv=sum(s['surplus_value_eur'] for s in yearly)
    td=sum(s['duration_min'] for s in yearly)
    
    ensure_dirs()
    txt = os.path.join(REPORT_DIR, f"ev_report_{year}.txt")
    with open(txt, 'w', encoding='utf-8') as f:
        f.write(f"{'='*45}\n  EV NABIJANIE - ROCNY REPORT {year}\n{'='*45}\n\n")
        f.write(f"  Sessions:    {len(yearly)}\n  Energia:     {te:.3f} kWh\n")
        f.write(f"    prebytky:  {ts:.3f} kWh ({ts/max(te,0.001)*100:.1f}%)\n")
        f.write(f"  Naklady:     {tc:.3f} EUR\n  Uspora:      {tv:.3f} EUR\n")
        f.write(f"  Cas:         {td:.0f} min ({td/60:.1f} h)\n")
        f.write(f"\n--- Mesiace ---\n")
        for m in sorted(by_month):
            d=by_month[m]
            f.write(f"  {m}: {d['n']}x, {d['e']:.3f} kWh (prebytky: {d['s']:.3f}), {d['c']:.3f} EUR, {d['d']:.0f} min\n")
        f.write(f"\n--- Pouzivatelia ---\n")
        for u, d in sorted(by_user.items()):
            f.write(f"  {u}: {d['n']}x, {d['e']:.3f} kWh (prebytky: {d['s']:.3f}), {d['c']:.3f} EUR\n")
    
    os.system(f"cp '{txt}' '{WWW_DIR}/ev_yearly_report.txt'; chmod 644 '{WWW_DIR}/ev_yearly_report.txt' '{txt}'")
    print(f"Yearly: {len(yearly)} sessions, {te:.3f} kWh, {tc:.3f} EUR")

def generate_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    sessions = read_csv()
    if not sessions: print("Ziadne data"); return
    
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sessions"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    sfill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    brd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    
    headers = ['Datum','Cas','Pouzivatel','Energia (kWh)','Naklady (EUR)','Tarifa',
               'Trvanie (min)','Dovod','Prebytky (kWh)','Siet (kWh)','Uspora (EUR)']
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=Alignment(horizontal='center'); cell.border=brd
    
    for ri,s in enumerate(sessions,2):
        dt=s.get('start_dt')
        vals=[dt.strftime('%Y-%m-%d') if dt else s.get('start_time',''), dt.strftime('%H:%M') if dt else '',
              s.get('user',''), s['energy_kwh'], s['cost_eur'], s.get('tariff',''),
              s['duration_min'], s.get('stop_reason',''), s['surplus_energy_kwh'], s['grid_energy_kwh'], s['surplus_value_eur']]
        for ci,v in enumerate(vals,1):
            cell=ws.cell(row=ri,column=ci,value=v); cell.border=brd
            if ci in (4,5,9,10,11): cell.number_format='0.000'
            if ci in (9,11) and isinstance(v,(int,float)) and v>0: cell.fill=sfill
    
    sr=len(sessions)+2
    ws.cell(row=sr,column=3,value="SPOLU:").font=Font(bold=True)
    for ci,k in [(4,'energy_kwh'),(5,'cost_eur'),(7,'duration_min'),(9,'surplus_energy_kwh'),(10,'grid_energy_kwh'),(11,'surplus_value_eur')]:
        c=ws.cell(row=sr,column=ci,value=round(sum(s[k] for s in sessions),3)); c.font=Font(bold=True)
    
    for i,w in enumerate([12,8,14,14,14,8,12,16,14,14,14]): ws.column_dimensions[chr(65+i)].width=w
    
    # Sheet 2: Mesacny
    ws2=wb.create_sheet("Mesacny prehlad")
    mh=['Mesiac','Sessions','Energia','Naklady','Prebytky','Siet','Uspora','Trvanie (min)']
    for c,h in enumerate(mh,1): cell=ws2.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill
    by_month=defaultdict(lambda:{k:0 for k in 'n e c s g v d'.split()})
    for s in sessions:
        if s.get('start_dt'):
            mk=s['start_dt'].strftime('%Y-%m')
            by_month[mk]['n']+=1; by_month[mk]['e']+=s['energy_kwh']; by_month[mk]['c']+=s['cost_eur']
            by_month[mk]['s']+=s['surplus_energy_kwh']; by_month[mk]['g']+=s['grid_energy_kwh']
            by_month[mk]['v']+=s['surplus_value_eur']; by_month[mk]['d']+=s['duration_min']
    for ri,(m,d) in enumerate(sorted(by_month.items()),2):
        for ci,v in enumerate([m,d['n'],round(d['e'],3),round(d['c'],3),round(d['s'],3),round(d['g'],3),round(d['v'],3),round(d['d'],0)],1):
            ws2.cell(row=ri,column=ci,value=v)
    
    # Sheet 3: Users
    ws3=wb.create_sheet("Pouzivatelia")
    uh=['Pouzivatel','Sessions','Energia','Naklady','Prebytky','Uspora']
    for c,h in enumerate(uh,1): cell=ws3.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill
    by_user=defaultdict(lambda:{k:0 for k in 'n e c s v'.split()})
    for s in sessions:
        u=s.get('user','?')
        by_user[u]['n']+=1; by_user[u]['e']+=s['energy_kwh']; by_user[u]['c']+=s['cost_eur']
        by_user[u]['s']+=s['surplus_energy_kwh']; by_user[u]['v']+=s['surplus_value_eur']
    for ri,(u,d) in enumerate(sorted(by_user.items()),2):
        for ci,v in enumerate([u,d['n'],round(d['e'],3),round(d['c'],3),round(d['s'],3),round(d['v'],3)],1):
            ws3.cell(row=ri,column=ci,value=v)
    
    ensure_dirs()
    p=os.path.join(WWW_DIR,"ev_charging_sessions.xlsx")
    wb.save(p); os.system(f"chmod 644 '{p}'")
    print(f"XLSX: {p} ({len(sessions)} sessions)")

def refresh_csv():
    if os.path.exists(CSV_PATH):
        os.makedirs(WWW_DIR, exist_ok=True)
        os.system(f"cp '{CSV_PATH}' '{WWW_DIR}/ev_charging_sessions.csv'; chmod 644 '{WWW_DIR}/ev_charging_sessions.csv'")
        print("CSV refreshed")
    else:
        print("CSV neexistuje")

if __name__ == "__main__":
    if len(sys.argv)<2: print("Pouzitie: monthly|yearly|xlsx|refresh"); sys.exit(1)
    cmd=sys.argv[1]; param=sys.argv[2] if len(sys.argv)>2 else None
    {'monthly':lambda:generate_monthly_report(param), 'yearly':lambda:generate_yearly_report(param),
     'xlsx':generate_xlsx, 'refresh':refresh_csv}.get(cmd, lambda:print(f"Neznamy: {cmd}"))()